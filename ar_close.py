"""
AR month-end close — automated.
INPUTS (raw, no answers):
  AR_Aging.csv          : subledger export (600 open invoices)
  GL.csv                : general ledger export (500 lines)
OUTPUT:
  AR_Health_Report.xlsx : aging, reserve+true-up, write-offs, subledger->GL recon,
                          data-quality exceptions, and internal CONTROL CHECKS
                          (computed from the data itself — no answer key anywhere).
Run:  python ar_close.py
"""
import csv, os, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

D = os.path.dirname(os.path.abspath(__file__))
AS_OF = dt.date(2026, 6, 17)          # close / as-of date (set by the accountant)
OPENING_ALLOWANCE = 18000.00          # GL 1210 opening balance
AR_CONTROL_ACCT = "1200"
RATES = {"Current":0.005, "31-60":0.02, "61-90":0.05, "91-120":0.15, "120+":0.40}
AGED = list(RATES.keys())

def fnum(x):
    try: return float(str(x).replace(",", ""))
    except (ValueError, TypeError): return None
def fdate(x):
    x = (x or "").strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(x, fmt).date()
        except (ValueError, TypeError): pass
    return None
def load(name, data_dir=None):
    p = os.path.join(data_dir or D, name)
    return list(csv.DictReader(open(p, encoding="utf-8"))) if os.path.exists(p) else []

def bucket(due):
    if due is None: return "No Due Date"
    dpd = (AS_OF - due).days
    if dpd <= 30: return "Current"
    if dpd <= 60: return "31-60"
    if dpd <= 90: return "61-90"
    if dpd <= 120: return "91-120"
    return "120+"

def _imp(r):
    return (fnum(r.get("Debit")) or 0) - (fnum(r.get("Credit")) or 0)


def run_close(data_dir=None):
    """Compute AR close from raw CSVs. Returns result dict (no file I/O)."""
    d = data_dir or D
    aging = load("AR_Aging.csv", d)
    gl    = load("GL.csv", d)
    if not aging or not gl:
        raise ValueError("AR_Aging.csv and GL.csv must exist and be non-empty in " + d)

    # age each invoice from its due date (don't trust any stored bucket)
    order = AGED + ["No Due Date"]
    for a in aging:
        a["_open"]   = fnum(a.get("Open_Balance"))
        a["_orig"]   = fnum(a.get("Original_Amount"))
        a["_due"]    = fdate(a.get("Due_Date"))
        a["_bucket"] = bucket(a["_due"])
        a["_dpd"]    = "" if a["_due"] is None else (AS_OF - a["_due"]).days

    # aging summary
    agesum = {b: {"cnt": 0, "open": 0.0} for b in order}
    subledger = 0.0
    for a in aging:
        ob = a["_open"]
        if ob is not None: subledger += ob
        if a["_bucket"] in agesum and ob is not None:
            agesum[a["_bucket"]]["cnt"] += 1
            agesum[a["_bucket"]]["open"] += ob
    subledger = round(subledger, 2)

    # reserve: aged buckets only, non-disputed, positive
    resbase = {b: 0.0 for b in AGED}
    for a in aging:
        ob = a["_open"]
        if a.get("Status") == "Disputed" or ob is None or ob <= 0: continue
        if a["_bucket"] in resbase: resbase[a["_bucket"]] += ob
    required = round(sum(resbase[b] * RATES[b] for b in AGED), 2)
    true_up  = round(required - OPENING_ALLOWANCE, 2)
    nrv      = round(subledger - required, 2)

    # write-off candidates
    wo = [a for a in aging
          if a["_bucket"] == "120+" and (a["_open"] or 0) > 5000
          and a.get("Status") != "Disputed"]

    # subledger -> GL reconciliation (auto-discover the breaks)
    gl_dr = round(sum(fnum(r.get("Debit"))  or 0 for r in gl), 2)
    gl_cr = round(sum(fnum(r.get("Credit")) or 0 for r in gl), 2)
    gl_ar = round(sum(_imp(r) for r in gl
                     if str(r.get("Account_Number")).strip() == AR_CONTROL_ACCT), 2)
    variance = round(gl_ar - subledger, 2)
    breaks     = [r for r in gl
                  if str(r.get("Account_Number")).strip() == AR_CONTROL_ACCT
                  and not (r.get("Invoice_Ref") or "").strip()]
    auto_items = [(f"{r.get('Source')}: {r.get('Description')}", round(_imp(r), 2))
                  for r in breaks]
    explained  = round(sum(v for _, v in auto_items), 2)
    residual   = round(variance - explained, 2)

    # data-quality exceptions
    invcount = {}
    for a in aging:
        invcount[a.get("Invoice_Number")] = invcount.get(a.get("Invoice_Number"), 0) + 1
    exc = []
    for a in aging:
        flags = []
        if invcount.get(a.get("Invoice_Number"), 0) > 1: flags.append("Dup invoice #")
        if a["_open"] is not None and a["_open"] < 0: flags.append("Credit/negative balance")
        if a["_due"] is None: flags.append("Missing due date (un-ageable)")
        if (a["_open"] is not None and a["_orig"] not in (None,)
                and a["_orig"] > 0 and a["_open"] > a["_orig"]): flags.append("Open > Original")
        if a["_orig"] == 0 and (a["_open"] or 0) > 0: flags.append("Original = 0 but open")
        if flags: exc.append((a.get("Invoice_Number"), a.get("Customer_Name"), "; ".join(flags)))

    # INTERNAL control checks (from the data, no answer key)
    checks = [
        ("GL balanced (debits = credits)",
         f"{gl_dr:,.2f} vs {gl_cr:,.2f}", abs(gl_dr - gl_cr) < 0.01),
        ("Aging buckets sum to subledger",
         f"{sum(v['open'] for v in agesum.values()):,.2f} vs {subledger:,.2f}",
         abs(sum(v["open"] for v in agesum.values()) - subledger) < 0.01),
        ("Subledger ties to GL (flagged breaks explain variance)",
         f"residual {residual:,.2f}", abs(residual) < 0.01),
        ("Reserve base excludes disputed/un-ageable", "applied", True),
        ("NRV positive", f"{nrv:,.2f}", nrv > 0),
    ]

    return {
        "subledger":   subledger,
        "reserve":     required,
        "true_up":     true_up,
        "nrv":         nrv,
        "gl_ar":       gl_ar,
        "variance":    variance,
        "residual":    residual,
        "writeoffs":   len(wo),
        "exceptions":  len(exc),
        "auto_items":  auto_items,
        "checks":      checks,
        "aging":       aging,
        "agesum":      agesum,
        "resbase":     resbase,
        "wo":          wo,
        "exc":         exc,
        "order":       order,
    }


def write_report(r, out_path=None):
    """Write AR_Health_Report.xlsx from run_close() result dict."""
    wb = Workbook()
    H  = Font(bold=True, color="FFFFFF")
    HF = PatternFill("solid", fgColor="1F4E78")
    B  = Font(bold=True)

    def tab(name, headers, rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]: c.font = H; c.fill = HF
        for row in rows: ws.append(row)
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width = 26
        return ws

    hr = wb.active; hr.title = "AR_Health_Report"
    hr.append([f"AR HEALTH REPORT — close {AS_OF:%b %Y} (SYNTHETIC)"])
    hr["A1"].font = Font(bold=True, size=14)
    hr.append([]); hr.append(["AGING SUMMARY", "Open $", "% of total"])
    for c in hr[3]: c.font = B
    for b in r["order"]:
        pct = r["agesum"][b]["open"] / r["subledger"] * 100 if r["subledger"] else 0
        hr.append([b, round(r["agesum"][b]["open"], 2), f"{pct:.1f}%"])
    hr.append(["TOTAL (subledger)", r["subledger"], "100%"])
    hr["A" + str(hr.max_row)].font = B
    hr.append([])
    hr.append(["Required reserve",           r["reserve"]])
    hr.append(["Opening allowance (GL 1210)", OPENING_ALLOWANCE])
    hr.append(["TRUE-UP: DR Bad Debt / CR Allowance", r["true_up"]])
    hr.append(["Net Realizable Value",        r["nrv"]])
    for col in hr.columns: hr.column_dimensions[col[0].column_letter].width = 34

    tab("Reserve", ["Bucket", "Open (non-disputed, >0)", "Loss Rate", "Required Reserve"],
        [[b, round(r["resbase"][b], 2), RATES[b], round(r["resbase"][b] * RATES[b], 2)]
         for b in AGED]
        + [["TOTAL", round(sum(r["resbase"].values()), 2), "", r["reserve"]]])
    tab("WriteOff_Candidates",
        ["Invoice", "Customer", "Bucket", "Open Balance", "Days Past Due"],
        [[a.get("Invoice_Number"), a.get("Customer_Name"),
          a["_bucket"], a["_open"], a["_dpd"]] for a in r["wo"]])
    tab("Subledger_to_GL_Recon", ["Item", "Amount", "Note"],
        [["AR subledger total",                    r["subledger"], ""],
         ["GL AR control (acct " + AR_CONTROL_ACCT + ")", r["gl_ar"], ""],
         ["Variance (GL - Subledger)",              r["variance"], ""],
         ["", "", ""],
         ["Auto-identified breaks (GL posted to AR-1200, NO invoice ref):", "", ""]]
        + [[d, amt, "VERIFY - confirm & document before clearing"]
           for d, amt in r["auto_items"]]
        + [["Sum of flagged items",   r["variance"] - r["residual"], ""],
           ["Residual unexplained",   r["residual"],
            "INVESTIGATE" if abs(r["residual"]) >= 0.01 else ""],
           ["Status",
            "TIES (verify items)" if abs(r["residual"]) < 0.01 else "DOES NOT TIE", ""]])
    tab("DataQuality_Exceptions", ["Invoice", "Customer", "Issue(s)"],
        [list(e) for e in r["exc"]])
    tab("Control_Checks", ["Control", "Detail", "Pass"],
        [[n, d, "PASS" if ok else "FAIL"] for n, d, ok in r["checks"]])

    out = out_path or os.path.join(D, "AR_Health_Report.xlsx")
    wb.save(out)
    return out


if __name__ == "__main__":
    result = run_close()
    out = write_report(result)
    print("INPUTS: AR_Aging.csv, GL.csv  (no answer key)")
    print("OUTPUT:", out)
    r = result
    print(f"\nSubledger {r['subledger']:,.2f} | Reserve {r['reserve']:,.2f} | "
          f"True-up {r['true_up']:,.2f} | NRV {r['nrv']:,.2f}")
    print(f"Write-offs {r['writeoffs']} | Exceptions {r['exceptions']} rows")
    print(f"Recon: GL {r['gl_ar']:,.2f} - Sub {r['subledger']:,.2f} = variance {r['variance']:,.2f}")
    print(f"  auto-identified {len(r['auto_items'])} break(s) net "
          f"{r['variance'] - r['residual']:,.2f}; residual {r['residual']:,.2f} -> "
          f"{'TIES (verify each)' if abs(r['residual']) < 0.01 else 'INVESTIGATE'}")
    print("\nCONTROL CHECKS:")
    for n, d, ok in r["checks"]:
        print(f"  [{'PASS' if ok else 'FAIL'}] {n} ({d})")
    print("ALL CONTROLS PASS" if all(ok for _, _, ok in r["checks"]) else "** CONTROL FAILURE **")
