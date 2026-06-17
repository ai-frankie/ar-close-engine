"""
AR CLOSE REVIEWER — a controls/QA check on a PREPARED workbook.
Recomputes the truth from raw data (AR_Aging.csv, GL.csv, reconciling_items.csv),
reads a human-prepared close workbook, and flags where the preparer's numbers are
wrong — and diagnoses WHY (sign flip, double-count, off-by).

Run:  python ar_review.py [path-to-prepared-workbook.xlsx]
If no path given (and no SAMPLE present), it builds AR_Close_Prepared_SAMPLE.xlsx
with two planted mistakes and reviews that — so it demos itself.
"""
import csv, os, sys, datetime as dt
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

D = os.path.dirname(os.path.abspath(__file__))
AS_OF = dt.date(2026, 6, 17); OPENING = 18000.0; ACCT = "1200"
RATES = {"Current":0.005,"31-60":0.02,"61-90":0.05,"91-120":0.15,"120+":0.40}

def fnum(x):
    try: return float(str(x).replace(",", ""))
    except: return None
def fdate(x):
    x=(x or "").strip()
    for f in ("%Y-%m-%d","%m/%d/%Y","%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(x,f).date()
        except: pass
    return None
def load(n):
    p=os.path.join(D,n); return list(csv.DictReader(open(p,encoding="utf-8"))) if os.path.exists(p) else []

# ---------- recompute TRUTH from raw data ----------
aging=load("AR_Aging.csv"); gl=load("GL.csv"); rec=load("reconciling_items.csv")
def bucket(due):
    if due is None: return "No Due Date"
    d=(AS_OF-due).days
    return "Current" if d<=30 else "31-60" if d<=60 else "61-90" if d<=90 else "91-120" if d<=120 else "120+"
subledger=round(sum(fnum(a.get("Open_Balance")) or 0 for a in aging),2)
resbase={b:0.0 for b in RATES}
for a in aging:
    ob=fnum(a.get("Open_Balance")); b=bucket(fdate(a.get("Due_Date")))
    if a.get("Status")=="Disputed" or ob is None or ob<=0: continue
    if b in resbase: resbase[b]+=ob
reserve=round(sum(resbase[b]*RATES[b] for b in RATES),2)
true_up=round(reserve-OPENING,2)
nrv=round(subledger-reserve,2)
gl_ar=round(sum((fnum(r.get("Debit")) or 0)-(fnum(r.get("Credit")) or 0)
            for r in gl if str(r.get("Account_Number")).strip()==ACCT),2)
variance=round(gl_ar-subledger,2)
writeoffs=sum(1 for a in aging if bucket(fdate(a.get("Due_Date")))=="120+"
              and (fnum(a.get("Open_Balance")) or 0)>5000 and a.get("Status")!="Disputed")
TRUTH={"subledger total":subledger,"gl ar control":gl_ar,"variance":variance,
       "required reserve":reserve,"true-up":true_up,"net realizable":nrv,"write-off":writeoffs}

# ---------- locate or build the prepared workbook ----------
path = sys.argv[1] if len(sys.argv)>1 else os.path.join(D,"AR_Close_Prepared_SAMPLE.xlsx")
if not os.path.exists(path):
    wb=Workbook(); ws=wb.active; ws.title="Close"
    # planted MISTAKES: subledger doubled, true-up added instead of subtracted
    for r in [["AR Close — Prepared",""],
              ["Subledger total",subledger*2],            # double-count error
              ["GL AR control acct 1200",gl_ar],
              ["Variance (GL - Sub)",gl_ar-subledger*2],
              ["Required reserve",reserve],
              ["Opening allowance",OPENING],
              ["True-up DR Bad Debt / CR Allowance",reserve+OPENING],  # sign flip
              ["Net Realizable Value",subledger-reserve],
              ["Write-off candidates",writeoffs]]:
        ws.append(r)
    wb.save(path)
    print("(no workbook given — built sample-with-errors:", os.path.basename(path), ")\n")

# ---------- read prepared values by LABEL (layout-robust) ----------
wb=load_workbook(path, data_only=True)
labelled=[]  # (label_lower, value)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        label=None; val=None
        for c in row:
            if c.value is None: continue
            if label is None and isinstance(c.value,str): label=c.value.strip().lower()
            elif val is None and isinstance(c.value,(int,float)): val=float(c.value)
        if label is not None and val is not None: labelled.append((label,val))
def reported(key):
    for lab,v in labelled:
        if key in lab: return v
    return None

# ---------- compare + diagnose ----------
rows=[]; fails=0
for key,exp in TRUTH.items():
    got=reported(key)
    if got is None:
        rows.append((key,"(missing)",f"{exp:,.2f}","MISSING","Metric not found in workbook")); fails+=1; continue
    ok=abs(got-exp)<=0.5
    diag=""
    if not ok:
        fails+=1
        if exp!=0 and abs(got/exp-2)<0.02: diag="~2x expected - double-counted (SUM hit a total row?)"
        elif key=="true-up" and abs(got-(reserve+OPENING))<0.5: diag="Added opening allowance instead of subtracting (sign flip)"
        elif abs(abs(got-exp)-2*OPENING)<0.5: diag="Off by 2× opening allowance — check +/- on the true-up"
        else: diag=f"Off by {got-exp:,.2f}"
    rows.append((key,f"{got:,.2f}",f"{exp:,.2f}","OK" if ok else "FAIL",diag))

# ---------- output ----------
out=Workbook(); ws=out.active; ws.title="Review"
ws.append(["Metric","Reported","Expected (recomputed)","Result","Diagnosis"])
for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78")
for r in rows: ws.append(list(r))
for col,w in zip("ABCDE",[20,18,22,10,46]): ws.column_dimensions[col].width=w
OUT=os.path.join(D,"AR_Close_Review.xlsx"); out.save(OUT)

print("AR CLOSE REVIEW —", os.path.basename(path))
print("OUTPUT:",OUT,"\n")
for m,g,e,res,diag in rows:
    print(f"  [{res}] {m}: reported {g} / expected {e}" + (f"  -> {diag}" if diag else ""))
print(f"\n{fails} issue(s) found." if fails else "\nClean — all reported figures tie to the recomputed truth.")
