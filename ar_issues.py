"""
AR ISSUES SCANNER — surfaces everything needing human attention before close.
Inputs (raw): AR_Aging.csv, GL.csv, reconciling_items.csv (analyst log).
Output: AR_Issues_Report.xlsx (one row per issue) + console summary.
Categories: DATA QUALITY | RECONCILIATION | WRITE-OFF | DISPUTE.
Run:  python ar_issues.py
"""
import csv, os, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

D = os.path.dirname(os.path.abspath(__file__))
AS_OF = dt.date(2026, 6, 17)
AR_CONTROL_ACCT = "1200"

def fnum(x):
    try: return float(str(x).replace(",", ""))
    except (ValueError, TypeError): return None
def fdate(x):
    x=(x or "").strip()
    for f in ("%Y-%m-%d","%m/%d/%Y","%Y-%m-%d %H:%M:%S"):
        try: return dt.datetime.strptime(x,f).date()
        except (ValueError, TypeError): pass
    return None
def load(n):
    p=os.path.join(D,n); return list(csv.DictReader(open(p,encoding="utf-8"))) if os.path.exists(p) else []

def bucket(due):
    if due is None: return "No Due Date"
    d=(AS_OF-due).days
    return "Current" if d<=30 else "31-60" if d<=60 else "61-90" if d<=90 else "91-120" if d<=120 else "120+"

def _imp(r): return (fnum(r.get("Debit")) or 0)-(fnum(r.get("Credit")) or 0)


def main():
    aging=load("AR_Aging.csv"); gl=load("GL.csv")
    issues=[]  # (Category, Severity, Ref, Customer, Issue, Amount, Action)

    # ---- DATA QUALITY ----
    invcount={}
    for a in aging: invcount[a.get("Invoice_Number")]=invcount.get(a.get("Invoice_Number"),0)+1
    for a in aging:
        inv=a.get("Invoice_Number"); cust=a.get("Customer_Name")
        ob=fnum(a.get("Open_Balance")); orig=fnum(a.get("Original_Amount")); due=fdate(a.get("Due_Date"))
        if invcount.get(inv,0)>1:
            issues.append(("DATA QUALITY","High",inv,cust,"Duplicate invoice number",ob,"De-dupe before posting"))
        if ob is not None and ob<0:
            issues.append(("DATA QUALITY","Medium",inv,cust,"Credit/negative balance",ob,"Confirm customer credit / reclass"))
        if due is None:
            issues.append(("DATA QUALITY","High",inv,cust,"Missing due date — cannot age",ob,"Get terms; re-age"))
        if ob is not None and orig is not None and orig>0 and ob>orig:
            issues.append(("DATA QUALITY","High",inv,cust,"Open balance exceeds original",ob,"Investigate over-application"))
        if orig==0 and (ob or 0)>0:
            issues.append(("DATA QUALITY","Medium",inv,cust,"Original amount = 0 but balance open",ob,"Fix invoice record"))

    # ---- RECONCILIATION (auto-discover breaks from GL) ----
    subledger=round(sum(fnum(a.get("Open_Balance")) or 0 for a in aging),2)
    gl_ar=round(sum(_imp(r) for r in gl if str(r.get("Account_Number")).strip()==AR_CONTROL_ACCT),2)
    variance=round(gl_ar-subledger,2)
    breaks=[r for r in gl if str(r.get("Account_Number")).strip()==AR_CONTROL_ACCT and not (r.get("Invoice_Ref") or "").strip()]
    explained=round(sum(_imp(r) for r in breaks),2)
    unexplained=round(variance-explained,2)
    for r in breaks:
        issues.append(("RECONCILIATION","High",r.get("Entry_ID") or "AR-1200","",
                       f"GL posting to AR with NO invoice ref - {r.get('Source')}: {r.get('Description')}",
                       round(_imp(r),2),"VERIFY: confirm cause, document, then clear"))
    if abs(unexplained)>=0.01:
        issues.append(("RECONCILIATION","High","AR-1200","","UNEXPLAINED variance subledger vs GL (residual after flagged items)",
                       unexplained,"Investigate - do not plug"))

    # ---- WRITE-OFF candidates (120+, >$5k, not disputed) ----
    for a in aging:
        ob=fnum(a.get("Open_Balance")); due=fdate(a.get("Due_Date"))
        if bucket(due)=="120+" and (ob or 0)>5000 and a.get("Status")!="Disputed":
            issues.append(("WRITE-OFF","High",a.get("Invoice_Number"),a.get("Customer_Name"),
                           "120+ days, >$5k — write-off candidate",ob,"Controller approval required"))

    # ---- DISPUTES ----
    for a in aging:
        if a.get("Status")=="Disputed":
            issues.append(("DISPUTE","Medium",a.get("Invoice_Number"),a.get("Customer_Name"),
                           "Disputed — excluded from reserve",fnum(a.get("Open_Balance")),"Resolve / document"))

    # ---- write report ----
    wb=Workbook(); ws=wb.active; ws.title="Issues"
    hdr=["Category","Severity","Invoice/Ref","Customer","Issue","Amount","Action"]
    ws.append(hdr)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="C00000")
    sev={"High":0,"Medium":1,"Low":2}
    for r in sorted(issues,key=lambda x:(x[0],sev.get(x[1],3))):
        ws.append(list(r))
    for col,w in zip("ABCDEFG",[16,10,14,26,42,16,30]): ws.column_dimensions[col].width=w
    OUT=os.path.join(D,"AR_Issues_Report.xlsx"); wb.save(OUT)

    # ---- console summary ----
    from collections import Counter
    cats=Counter(i[0] for i in issues); sevs=Counter(i[1] for i in issues)
    print("AR ISSUES SCAN —", f"{AS_OF:%b %Y}")
    print("OUTPUT:",OUT)
    print(f"\nTotal issues: {len(issues)}")
    for c in ["DATA QUALITY","RECONCILIATION","WRITE-OFF","DISPUTE"]:
        if cats.get(c): print(f"  {c:16} {cats[c]}")
    print(f"Severity: High {sevs.get('High',0)} | Medium {sevs.get('Medium',0)} | Low {sevs.get('Low',0)}")
    print(f"\nSubledger {subledger:,.2f} vs GL {gl_ar:,.2f} -> variance {variance:,.2f}; auto-flagged breaks net {explained:,.2f}; residual {unexplained:,.2f}")
    print("Top issues (first 8):")
    for r in sorted(issues,key=lambda x:(sev.get(x[1],3),x[0]))[:8]:
        amt=f"{r[5]:,.2f}" if isinstance(r[5],(int,float)) else ""
        print(f"  [{r[1]}] {r[0]} | {r[2]} {r[3]} | {r[4]} {amt}")


if __name__ == "__main__":
    main()
